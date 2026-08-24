# ============================================================
# Atmos Energy Invoice Folder Extractor -> Excel Summary
# CORRECTED MULTI-YEAR VERSION FOR 2019-2026 PDFs + INCREMENTAL EXCEL UPDATE
#
# Key fixes:
#   1. Reads only PDF files with "Invoice" in the filename.
#   2. Ignores GCR rate / filing PDFs.
#   3. Recognizes Sales Tax lines found in the May 2026 invoice.
#   4. Writes any extra / unexpected charge type to end columns.
#   5. Flags invoices with extra charge types as NEEDS REVIEW.
#   6. Orders Summary, Raw Line Items, and Validation by billing month.
#   7. Parses invoice line rows from text first, not fragile PDF block order.
#   8. Does NOT force a fixed 15-line default description list.
#
# Install first:
#   pip install pymupdf openpyxl
#
# Run:
#   python USE_THIS_extract_atmos_invoice_FINAL_EXTRAS_REVIEW.py
# ============================================================

import re
import argparse
import calendar
from pathlib import Path
from datetime import datetime, date
from copy import copy

import fitz  # PyMuPDF
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# PORTABLE PROJECT DEFAULTS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parent

# By default, gas invoices live under year folders such as:
# data/gas/input/2019, data/gas/input/2020, ..., data/gas/input/2026
DEFAULT_INPUT_FOLDER = PROJECT_ROOT / "data" / "gas" / "input"
DEFAULT_OUTPUT_XLSX = PROJECT_ROOT / "output" / "gas" / "atmos_energy_invoices.xlsx"

# Historical default range. Override from the command line with --years.
DEFAULT_YEAR_FOLDERS = [str(year) for year in range(2019, 2027)]

PDF_FILE_PATTERN = "*.pdf"


# ============================================================
# OUTPUT COLUMNS
# ============================================================

SUMMARY_COLUMNS = [
    "Month",
    "Service Account",
    "Customer Name",
    "Facility ID",
    "Address",
    "City",
    "State",
    "Zip",
    "Current Charges",
    "Billed MCF",
    "Billed MMBtu",
    "1st 0 to 1,500 MMBtu Rate",
    "Next 3,500 MMBtu Rate",
    "All MMBtu over 5,000 MMBtu Rate",
    "1st 0 to 1,500 MMBtu Amount",
    "Next 3,500 MMBtu Amount",
    "All MMBtu over 5,000 MMBtu Amount",
    "Total Cost Based on MMBtu",
    "Customer Charge",
    "Plant Protection Fee",
    "Reimbursement of MGRT 1",
    "Reimbursement of MGRT 2",
    "Reimbursement of MGRT 3",
    "Reimbursement of MGRT 4",
    "Total Reimbursement of MGRT",
    "Street & Alley Fees 1",
    "Street & Alley Fees 2",
    "Street & Alley Fees 3",
    "Street & Alley Fees 4",
    "Total Street & Alley Fees",
    "FIN 48 Tax Refund",
    "Pipeline Safety Fee",
    "GCR - Industrial Sales ($/MMBtu)",
    "GCR - Transportation ($/MMBtu)",
    "Billed CCF",
    "Calculated Rate ($/MMBTU)",
    "From Billing Date",
    "To Billing Date",
    "Extra Charge Details",
    "Extra Charge Total",
    "Review Status",
    "Review Notes",
]

DESCRIPTION_NAMES = [
    "Customer Charge",
    "Reimbursement of MGRT",
    "Street & Alley Fee",
    "Pipeline Safety Fee",
    "FIN 48 Tax Refund",
    "Gas Cost Recovery",
    "Sales Service Rate",
    "Plant Protection Fee",
    "Sales Tax",
]

DESCRIPTION_PATTERN = (
    r"Customer\s+Charge|"
    r"Reimbursement\s+of\s+MGRT|"
    r"Street\s*&\s*Alley\s+Fee|"
    r"Pipeline\s+Safety\s+Fee|"
    r"FIN\s+48\s+Tax\s+Refund|"
    r"Gas\s+Cost\s+Recovery|"
    r"Sales\s+Service\s+Rate|"
    r"Plant\s+Protection\s+Fee|"
    r"Sales\s+Tax"
)

# These descriptions are expected in the standard summary layout.
# Anything outside this list is written to the end columns and marked NEEDS REVIEW.
STANDARD_SUMMARY_DESCRIPTIONS = {
    "Customer Charge",
    "Reimbursement of MGRT",
    "Street & Alley Fee",
    "Pipeline Safety Fee",
    "FIN 48 Tax Refund",
    "Gas Cost Recovery",
    "Sales Service Rate",
    "Plant Protection Fee",
}

# Used only if the PDF text does not expose the description lines at all.
# Do not use this when the invoice has real descriptions available.
DEFAULT_DESCRIPTION_ORDER = [
    "Customer Charge",
    "Reimbursement of MGRT",
    "Street & Alley Fee",
    "Reimbursement of MGRT",
    "Street & Alley Fee",
    "FIN 48 Tax Refund",
    "Reimbursement of MGRT",
    "Street & Alley Fee",
    "Gas Cost Recovery",
    "Sales Service Rate",
    "Sales Service Rate",
    "Sales Service Rate",
    "Reimbursement of MGRT",
    "Street & Alley Fee",
]


# ============================================================
# CONVERTERS
# ============================================================


def money_to_float(value):
    if value is None:
        return 0.0

    value = str(value).strip().replace("\u00a0", "")
    negative = value.startswith("(") and value.endswith(")")

    clean = (
        value.replace("$", "")
        .replace(",", "")
        .replace("(", "")
        .replace(")", "")
        .strip()
    )

    if clean == "":
        return 0.0

    number = float(clean)
    return -number if negative else number


def number_to_float(value):
    if value is None:
        return 0.0

    clean = str(value).strip().replace("\u00a0", "").replace(",", "")

    if clean == "":
        return 0.0

    return float(clean)


def clean_lines(text):
    return [line.strip() for line in text.splitlines() if line.strip()]


# ============================================================
# PDF READERS
# ============================================================


def extract_pdf_text(pdf_path):
    pdf_path = Path(pdf_path)

    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF file not found:\n{pdf_path}")

    text_parts = []

    with fitz.open(pdf_path) as doc:
        for page in doc:
            text_parts.append(page.get_text("text"))

    return "\n".join(text_parts)


def extract_pdf_words(pdf_path):
    words = []

    with fitz.open(pdf_path) as doc:
        for page_index, page in enumerate(doc):
            for w in page.get_text("words"):
                x0, y0, x1, y1, word, block_no, line_no, word_no = w

                words.append(
                    {
                        "page": page_index,
                        "x0": x0,
                        "y0": y0,
                        "x1": x1,
                        "y1": y1,
                        "word": word.strip(),
                        "block": block_no,
                        "line": line_no,
                        "word_no": word_no,
                    }
                )

    return words


# ============================================================
# DESCRIPTION / CUSTOMER PARSING
# ============================================================


def standardize_description(description):
    description = re.sub(r"\s+", " ", description).strip()

    for standard_name in DESCRIPTION_NAMES:
        if description.lower() == standard_name.lower():
            return standard_name

    return description


def parse_descriptions_and_customer(text):
    descriptions = []
    service_account = ""
    customer_name = ""

    line_pattern = re.compile(
        rf"^(\d{{10}})\s+(.+?)\s+({DESCRIPTION_PATTERN})$",
        re.IGNORECASE,
    )

    for line in clean_lines(text):
        match = line_pattern.match(line)

        if not match:
            continue

        service_account = match.group(1).strip()
        customer_name = match.group(2).strip()
        description = standardize_description(match.group(3))
        descriptions.append(description)

    if not service_account:
        account_match = re.search(r"\b(080\d{7})\b", text)
        service_account = account_match.group(1) if account_match else ""

    if not customer_name:
        customer_name = ""

    # Important: use the invoice's real description lines when present.
    # Do NOT force a 15-line list, because May has Sales Tax and April has 14 lines.
    if not descriptions:
        descriptions = DEFAULT_DESCRIPTION_ORDER[:]

    return descriptions, service_account, customer_name


# ============================================================
# LINE ITEM PARSING
# ============================================================


def _build_line_item(line_no, prod_date, numeric_tokens, amount_token, descriptions, service_account, customer_name):
    mcf = 0.0
    mmbtu = 0.0
    rate = 0.0

    if len(numeric_tokens) >= 3:
        mcf = number_to_float(numeric_tokens[-3])
        mmbtu = number_to_float(numeric_tokens[-2])
        rate = number_to_float(numeric_tokens[-1])
    elif len(numeric_tokens) == 2:
        mcf = number_to_float(numeric_tokens[0])
        mmbtu = number_to_float(numeric_tokens[1])
    elif len(numeric_tokens) == 1:
        rate = number_to_float(numeric_tokens[0])

    description = ""
    if line_no - 1 < len(descriptions):
        description = descriptions[line_no - 1]

    return {
        "Line #": line_no,
        "Prod Date": prod_date,
        "Service Account": service_account,
        "Customer Name": customer_name,
        "Description": description,
        "MCF": mcf,
        "MMBtu": mmbtu,
        "Rate": rate,
        "Amount": money_to_float(amount_token),
        "Amount Text": amount_token,
    }


def parse_line_items_from_text_rows(text):
    """Parse rows like:
       11 May-26 28,786 28,859 4.87240 $140,612.59
       5 May-26 ($403.02)

       This is more reliable for these Atmos PDFs than block-position parsing.
    """
    descriptions, service_account, customer_name = parse_descriptions_and_customer(text)

    row_re = re.compile(
        r"^\s*(\d{1,3})\s+([A-Za-z]{3}-\d{2})(?:\s+(.*?))?\s+(\(?\$[\d,]+\.\d{2}\)?)\s*$"
    )

    number_re = re.compile(r"^[\d,]+(?:\.\d+)?$")

    by_line_no = {}

    for line in clean_lines(text):
        match = row_re.match(line)

        if not match:
            continue

        line_no = int(match.group(1))
        prod_date = match.group(2)
        middle = match.group(3) or ""
        amount_token = match.group(4)

        if not (1 <= line_no <= 999):
            continue

        middle_tokens = middle.split()
        numeric_tokens = [token for token in middle_tokens if number_re.match(token)]

        by_line_no[line_no] = _build_line_item(
            line_no=line_no,
            prod_date=prod_date,
            numeric_tokens=numeric_tokens,
            amount_token=amount_token,
            descriptions=descriptions,
            service_account=service_account,
            customer_name=customer_name,
        )

    return [by_line_no[key] for key in sorted(by_line_no)]


def parse_line_items_from_words(words, text):
    """Parse the visible Atmos invoice table using word coordinates.

    Some invoice PDFs do not return one clean text line per
    invoice row. Word coordinates are more reliable because each row has stable
    columns: line number, prod date, account, description, MCF, MMBtu, rate,
    amount.
    """
    _, fallback_service_account, fallback_customer_name = parse_descriptions_and_customer(text)

    month_re = re.compile(r"^[A-Za-z]{3}-\d{2}$")
    line_no_re = re.compile(r"^\d{1,3}$")
    amount_re = re.compile(r"^\(?\$[\d,]+\.\d{2}\)?$")
    number_re = re.compile(r"^[\d,]+(?:\.\d+)?$")

    line_items = []

    pages = sorted({word["page"] for word in words})

    for page_no in pages:
        page_words = [word for word in words if word["page"] == page_no]

        # Keep only the body area where invoice line items appear.
        table_words = [
            word
            for word in page_words
            if 180 <= word["y0"] <= 470
        ]

        table_words.sort(key=lambda word: (word["y0"], word["x0"]))

        rows = []
        for word in table_words:
            if not rows:
                rows.append([word])
                continue

            current_row = rows[-1]
            avg_y = sum(item["y0"] for item in current_row) / len(current_row)

            if abs(word["y0"] - avg_y) <= 3.0:
                current_row.append(word)
            else:
                rows.append([word])

        for row_words in rows:
            row_words = sorted(row_words, key=lambda word: word["x0"])

            line_no = None
            prod_date = ""
            service_account = fallback_service_account
            customer_name = fallback_customer_name
            description_words = []
            amount_token = None
            numeric_tokens = []

            for word in row_words:
                token = word["word"]
                x0 = word["x0"]

                if x0 < 40 and line_no_re.match(token):
                    line_no = int(token)
                elif 15 <= x0 < 75 and month_re.match(token):
                    prod_date = token
                elif 60 <= x0 < 160 and re.match(r"^080\d{7}$", token):
                    service_account = token
                elif 175 <= x0 < 340:
                    # Keep the displayed customer name if present.
                    pass
                elif 350 <= x0 < 520:
                    description_words.append(token)
                elif 520 <= x0 < 700 and number_re.match(token):
                    numeric_tokens.append((x0, token))
                elif x0 >= 690 and amount_re.match(token):
                    amount_token = token

            if line_no is None or not prod_date or amount_token is None:
                continue

            description = standardize_description(" ".join(description_words))

            numeric_tokens = [token for _, token in sorted(numeric_tokens, key=lambda item: item[0])]

            mcf = 0.0
            mmbtu = 0.0
            rate = 0.0

            if len(numeric_tokens) >= 3:
                mcf = number_to_float(numeric_tokens[-3])
                mmbtu = number_to_float(numeric_tokens[-2])
                rate = number_to_float(numeric_tokens[-1])
            elif len(numeric_tokens) == 2:
                mcf = number_to_float(numeric_tokens[0])
                mmbtu = number_to_float(numeric_tokens[1])
            elif len(numeric_tokens) == 1:
                rate = number_to_float(numeric_tokens[0])

            line_items.append(
                {
                    "Line #": line_no,
                    "Prod Date": prod_date,
                    "Service Account": service_account,
                    "Customer Name": customer_name,
                    "Description": description,
                    "MCF": mcf,
                    "MMBtu": mmbtu,
                    "Rate": rate,
                    "Amount": money_to_float(amount_token),
                    "Amount Text": amount_token,
                }
            )

    by_line_no = {}
    for item in line_items:
        # Deduplicate duplicate pages by line number. The April PDF exposes duplicate pages.
        by_line_no[item["Line #"]] = item

    return [by_line_no[key] for key in sorted(by_line_no)]


def parse_line_items_from_text_fallback(text):
    descriptions, service_account, customer_name = parse_descriptions_and_customer(text)

    tokens = re.findall(
        r"\(?\$[\d,]+\.\d{2}\)?|" r"[A-Za-z]{3}-\d{2}|" r"[\d,]+\.\d+|" r"[\d,]+",
        text,
    )

    month_re = re.compile(r"^[A-Za-z]{3}-\d{2}$")
    amount_re = re.compile(r"^\(?\$[\d,]+\.\d{2}\)?$")
    line_no_re = re.compile(r"^\d{1,3}$")
    number_re = re.compile(r"^[\d,]+(?:\.\d+)?$")

    by_line_no = {}
    i = 0

    while i < len(tokens) - 2:
        if line_no_re.match(tokens[i]) and month_re.match(tokens[i + 1]):
            line_no = int(tokens[i])
            prod_date = tokens[i + 1]

            j = i + 2
            numeric_tokens = []
            amount_token = None

            while j < len(tokens):
                if amount_re.match(tokens[j]):
                    amount_token = tokens[j]
                    break

                if number_re.match(tokens[j]):
                    numeric_tokens.append(tokens[j])

                j += 1

            if amount_token is not None and 1 <= line_no <= 999:
                by_line_no[line_no] = _build_line_item(
                    line_no=line_no,
                    prod_date=prod_date,
                    numeric_tokens=numeric_tokens,
                    amount_token=amount_token,
                    descriptions=descriptions,
                    service_account=service_account,
                    customer_name=customer_name,
                )

                i = j + 1
                continue

        i += 1

    return [by_line_no[key] for key in sorted(by_line_no)]


def parse_invoice_lines(pdf_path, text):
    words = extract_pdf_words(pdf_path)

    # Preferred parser for invoice PDFs that preserve stable word coordinates.
    line_items = parse_line_items_from_words(words, text)

    # Text-row fallback for PDFs that extract rows in normal reading order.
    if len(line_items) < 5:
        line_items = parse_line_items_from_text_rows(text)

    # Final token fallback. Reject tiny results because those are usually
    # header dates/due dates, not invoice detail rows.
    if len(line_items) < 5:
        line_items = parse_line_items_from_text_fallback(text)

    if len(line_items) < 5:
        debug_path = Path(pdf_path).with_name(f"DEBUG_{Path(pdf_path).stem}_text.txt")
        debug_path.write_text(text, encoding="utf-8", errors="replace")

        raise ValueError(
            "No valid invoice line items were found. " f"Debug text saved here: {debug_path}"
        )

    return line_items

# ============================================================
# BILL TO / DATE HELPERS
# ============================================================


def extract_bill_to_info(text):
    lines = clean_lines(text)

    bill_lines = []
    capture = False

    for line in lines:
        if line.upper() == "BILL TO:":
            capture = True
            continue

        if capture and line.upper() == "CHECK REMITTANCE TO:":
            break

        if capture:
            bill_lines.append(line)

    address = ""
    city = ""
    state = ""
    zip_code = ""

    for line in bill_lines:
        if re.match(r"^\d+\s+", line):
            address = line.strip()

        city_match = re.match(r"^(.+),\s*([A-Z]{2})\s+(\d{5}(?:-\d{4})?)$", line)

        if city_match:
            city = city_match.group(1).strip()
            state = city_match.group(2).strip()
            zip_code = city_match.group(3).strip()

    return address, city, state, zip_code


def find_first(pattern, text, default=""):
    match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
    return match.group(1).strip() if match else default


def get_month_dates(prod_date):
    dt = datetime.strptime(prod_date, "%b-%y")

    first_day = date(dt.year, dt.month, 1)
    last_day = date(dt.year, dt.month, calendar.monthrange(dt.year, dt.month)[1])

    return prod_date, first_day, last_day


def month_sort_date(month_value):
    """Convert Apr-26 / May-26 to a real date for month ordering."""
    try:
        dt = datetime.strptime(str(month_value).strip(), "%b-%y")
        return date(dt.year, dt.month, 1)
    except Exception:
        return date.max


def line_item_sort_key(item):
    """Sort raw line items by billing month, then source PDF, then line number."""
    return (
        month_sort_date(item.get("Prod Date", "")),
        str(item.get("Source PDF", "")).lower(),
        int(item.get("Line #", 0) or 0),
    )


# ============================================================
# SUMMARY CALCULATION
# ============================================================


def amount_sum(line_items, description):
    return sum(
        item["Amount"]
        for item in line_items
        if item["Description"].lower() == description.lower()
    )


def find_by_description(line_items, description):
    return [
        item
        for item in line_items
        if item["Description"].lower() == description.lower()
    ]


def get_item_amount(items, index):
    return items[index]["Amount"] if len(items) > index else 0.0


def get_item_rate(items, index):
    return items[index]["Rate"] if len(items) > index else 0.0


def find_extra_items(line_items):
    """Return charge rows that are not part of the normal summary layout.

    Example: the May 2026 invoice includes Sales Tax rows. Those are valid
    invoice rows, but they are different from the regular April layout, so
    they are sent to the end review columns instead of being silently ignored.
    """
    extras = []

    for item in line_items:
        description = (item.get("Description") or "").strip()

        if not description or description not in STANDARD_SUMMARY_DESCRIPTIONS:
            extras.append(item)

    return extras


def format_money(value):
    if value < 0:
        return f"(${abs(value):,.2f})"
    return f"${value:,.2f}"


def format_extra_charge_details(extra_items):
    if not extra_items:
        return ""

    detail_parts = []

    for item in extra_items:
        description = item.get("Description") or "UNKNOWN DESCRIPTION"
        detail_parts.append(
            f"Line {item['Line #']} - {description}: {format_money(item['Amount'])}"
        )

    return "; ".join(detail_parts)


def build_review_status_and_notes(extra_items):
    if not extra_items:
        return "OK", ""

    descriptions = sorted({(item.get("Description") or "UNKNOWN DESCRIPTION") for item in extra_items})
    review_notes = "Extra/different charge type found: " + ", ".join(descriptions)

    return "NEEDS REVIEW", review_notes


def build_summary_row(text, line_items):
    first_item = line_items[0]

    month_value, from_billing_date, to_billing_date = get_month_dates(
        first_item["Prod Date"]
    )

    service_account = first_item["Service Account"]
    customer_name = first_item["Customer Name"]

    facility_id = find_first(r"\b(\d{5}-DS)\b", text)

    address, city, state, zip_code = extract_bill_to_info(text)

    gas_items = find_by_description(line_items, "Gas Cost Recovery")
    sales_items = find_by_description(line_items, "Sales Service Rate")
    mgrt_items = find_by_description(line_items, "Reimbursement of MGRT")
    street_items = find_by_description(line_items, "Street & Alley Fee")
    extra_items = find_extra_items(line_items)

    gas_item = gas_items[0] if gas_items else None

    billed_mcf = gas_item["MCF"] if gas_item else 0.0
    billed_mmbtu = gas_item["MMBtu"] if gas_item else 0.0

    tier1_rate = get_item_rate(sales_items, 0)
    tier2_rate = get_item_rate(sales_items, 1)
    tier3_rate = get_item_rate(sales_items, 2)

    tier1_amount = get_item_amount(sales_items, 0)
    tier2_amount = get_item_amount(sales_items, 1)
    tier3_amount = get_item_amount(sales_items, 2)

    total_cost_based_on_mmbtu = tier1_amount + tier2_amount + tier3_amount

    current_charges = sum(item["Amount"] for item in line_items)

    customer_charge = amount_sum(line_items, "Customer Charge")
    plant_protection_fee = amount_sum(line_items, "Plant Protection Fee")

    mgrt_1 = get_item_amount(mgrt_items, 0)
    mgrt_2 = get_item_amount(mgrt_items, 1)
    mgrt_3 = get_item_amount(mgrt_items, 2)
    mgrt_4 = get_item_amount(mgrt_items, 3)
    total_mgrt = mgrt_1 + mgrt_2 + mgrt_3 + mgrt_4

    street_1 = get_item_amount(street_items, 0)
    street_2 = get_item_amount(street_items, 1)
    street_3 = get_item_amount(street_items, 2)
    street_4 = get_item_amount(street_items, 3)
    total_street = street_1 + street_2 + street_3 + street_4

    fin_48_tax_refund = amount_sum(line_items, "FIN 48 Tax Refund")
    pipeline_safety_fee = amount_sum(line_items, "Pipeline Safety Fee")

    gcr_industrial_sales_rate = gas_item["Rate"] if gas_item else 0.0

    # Keep this column blank unless you later connect it to the monthly GCR rate filing.
    gcr_transportation_rate = None

    billed_ccf = billed_mcf * 10
    calculated_rate = current_charges / billed_mmbtu if billed_mmbtu else 0.0

    extra_charge_details = format_extra_charge_details(extra_items)
    extra_charge_total = sum(item["Amount"] for item in extra_items)
    review_status, review_notes = build_review_status_and_notes(extra_items)

    return [
        month_value,
        service_account,
        customer_name,
        facility_id,
        address,
        city,
        state,
        zip_code,
        current_charges,
        billed_mcf,
        billed_mmbtu,
        tier1_rate,
        tier2_rate,
        tier3_rate,
        tier1_amount,
        tier2_amount,
        tier3_amount,
        total_cost_based_on_mmbtu,
        customer_charge,
        plant_protection_fee,
        mgrt_1,
        mgrt_2,
        mgrt_3,
        mgrt_4,
        total_mgrt,
        street_1,
        street_2,
        street_3,
        street_4,
        total_street,
        fin_48_tax_refund,
        pipeline_safety_fee,
        gcr_industrial_sales_rate,
        gcr_transportation_rate,
        billed_ccf,
        calculated_rate,
        from_billing_date,
        to_billing_date,
        extra_charge_details,
        extra_charge_total,
        review_status,
        review_notes,
    ]



# ============================================================
# EXISTING EXCEL / INCREMENTAL UPDATE HELPERS
# ============================================================


def normalize_source_label(value):
    """Normalize Source PDF labels so 2026\file.pdf and 2026/file.pdf match."""
    return str(value or "").strip().replace("/", "\\").lower()


def source_filename_only(value):
    """Return only the PDF filename from a source label."""
    normalized = normalize_source_label(value)
    return normalized.split("\\")[-1]


def coerce_month_start(value):
    """Convert Excel values such as Apr-26 or 4/1/2026 to first day of month."""
    if value is None:
        return None

    if isinstance(value, datetime):
        return date(value.year, value.month, 1)

    if isinstance(value, date):
        return date(value.year, value.month, 1)

    text_value = str(value).strip()
    if not text_value:
        return None

    for fmt in ("%b-%y", "%b-%Y", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(text_value, fmt)
            return date(dt.year, dt.month, 1)
        except ValueError:
            pass

    return None


def get_header_map(ws):
    """Return {header name: 1-based column number}."""
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None and str(cell.value).strip()
    }


def read_existing_excel_state(output_path):
    """Read existing workbook and find latest saved billing month + existing PDFs.

    The latest month is read from Gas Invoice Summary using From Billing Date.
    Existing PDF labels are read from Validation / Source PDF to avoid duplicates.
    """
    output_path = Path(output_path)

    if not output_path.exists():
        return None, set(), set()

    wb = load_workbook(output_path, data_only=True)

    latest_month = None
    existing_sources = set()
    legacy_existing_source_names = set()

    if "Gas Invoice Summary" in wb.sheetnames:
        ws = wb["Gas Invoice Summary"]
        header_map = get_header_map(ws)

        month_col = header_map.get("From Billing Date") or header_map.get("Month")

        if month_col:
            for row_idx in range(2, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=month_col).value
                month_value = coerce_month_start(value)

                if month_value and (latest_month is None or month_value > latest_month):
                    latest_month = month_value

    if "Validation" in wb.sheetnames:
        ws = wb["Validation"]
        header_map = get_header_map(ws)
        source_col = header_map.get("Source PDF")

        if source_col:
            for row_idx in range(2, ws.max_row + 1):
                source_value = ws.cell(row=row_idx, column=source_col).value

                if source_value:
                    normalized_source = normalize_source_label(source_value)
                    existing_sources.add(normalized_source)

                    # Older versions of this script saved only the PDF filename,
                    # not the year-relative path. Keep filename-only matching only
                    # for those legacy rows so same filenames in different year
                    # folders are not skipped incorrectly.
                    if "\\" not in normalized_source:
                        legacy_existing_source_names.add(source_filename_only(source_value))

    wb.close()
    return latest_month, existing_sources, legacy_existing_source_names


def apply_row_style_from_previous(ws, row_idx):
    """Copy formatting from the previous row to a newly appended row."""
    if row_idx <= 2:
        return

    previous_row_idx = row_idx - 1

    for col_idx in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=previous_row_idx, column=col_idx)
        target_cell = ws.cell(row=row_idx, column=col_idx)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)


def set_status_fill(cell, ok_values=("OK",)):
    """Color a status cell green for OK values, red otherwise."""
    cell.font = Font(bold=True)

    if str(cell.value or "").upper() in ok_values:
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
    else:
        cell.fill = PatternFill("solid", fgColor="FFC7CE")


def append_summary_rows(ws, summary_rows):
    if not summary_rows:
        return

    review_col_idx = SUMMARY_COLUMNS.index("Review Status") + 1

    for summary_row in summary_rows:
        row_idx = ws.max_row + 1
        ws.append(summary_row)
        apply_row_style_from_previous(ws, row_idx)
        set_status_fill(ws.cell(row=row_idx, column=review_col_idx))


def append_raw_line_items(ws, all_line_items):
    if not all_line_items:
        return

    for item in all_line_items:
        row_idx = ws.max_row + 1
        description = (item.get("Description") or "").strip()
        review_flag = "OK" if description and description in STANDARD_SUMMARY_DESCRIPTIONS else "NEEDS REVIEW"

        ws.append(
            [
                item.get("Source PDF", ""),
                item["Line #"],
                item["Prod Date"],
                item["Service Account"],
                item["Customer Name"],
                item["Description"],
                item["MCF"],
                item["MMBtu"],
                item["Rate"],
                item["Amount"],
                review_flag,
            ]
        )

        apply_row_style_from_previous(ws, row_idx)
        set_status_fill(ws.cell(row=row_idx, column=11))


def append_validation_rows(ws, validation_rows):
    if not validation_rows:
        return

    for row_data in validation_rows:
        row_idx = ws.max_row + 1
        ws.append(row_data)
        apply_row_style_from_previous(ws, row_idx)
        set_status_fill(ws.cell(row=row_idx, column=7))
        set_status_fill(ws.cell(row=row_idx, column=8))


def append_failed_files_rows(wb, failed_files):
    if not failed_files:
        return

    if "Failed Files" in wb.sheetnames:
        ws = wb["Failed Files"]
    else:
        ws = wb.create_sheet("Failed Files")
        write_failed_files_sheet(ws, [])

    for row_data in failed_files:
        row_idx = ws.max_row + 1
        ws.append(row_data)
        apply_row_style_from_previous(ws, row_idx)


def append_to_existing_excel(output_path, summary_rows, all_line_items, validation_rows, failed_files):
    """Append newly processed invoices into the existing Excel workbook."""
    output_path = Path(output_path)

    wb = load_workbook(output_path)

    required_sheets = {"Gas Invoice Summary", "Raw Line Items", "Validation"}
    missing_sheets = [name for name in required_sheets if name not in wb.sheetnames]

    if missing_sheets:
        wb.close()
        raise ValueError(
            "Existing Excel file is missing required sheet(s): " + ", ".join(missing_sheets)
        )

    append_summary_rows(wb["Gas Invoice Summary"], summary_rows)
    append_raw_line_items(wb["Raw Line Items"], all_line_items)
    append_validation_rows(wb["Validation"], validation_rows)
    append_failed_files_rows(wb, failed_files)

    wb.save(output_path)
    wb.close()


def process_new_invoices(base_folder, year_folders, latest_month, existing_sources, legacy_existing_source_names):
    """Process only PDFs newer than the latest month already in Excel.

    This still parses candidate PDFs so the decision is based on invoice billing
    month, not filename date guesses.
    """
    pdf_files = get_pdf_files(base_folder, year_folders)
    base_folder_path = Path(base_folder)

    if not pdf_files:
        years_text = ", ".join(year_folders or DEFAULT_YEAR_FOLDERS)
        raise FileNotFoundError(
            f"No invoice PDF files found in year folders {years_text} under:\n{base_folder}"
        )

    summary_rows = []
    all_line_items = []
    validation_rows = []
    failed_files = []

    skipped_existing = 0
    skipped_old_or_same_month = 0

    for pdf_path in pdf_files:
        try:
            source_label = str(pdf_path.relative_to(base_folder_path))
        except ValueError:
            source_label = pdf_path.name

        normalized_source = normalize_source_label(source_label)
        normalized_name = source_filename_only(source_label)

        if normalized_source in existing_sources or normalized_name in legacy_existing_source_names:
            skipped_existing += 1
            continue

        try:
            summary_row, line_items = process_single_pdf(pdf_path)
            invoice_month = coerce_month_start(summary_row[SUMMARY_COLUMNS.index("From Billing Date")])

            if latest_month and invoice_month and invoice_month <= latest_month:
                skipped_old_or_same_month += 1
                print(
                    f"SKIP: {source_label} | {summary_row[0]} is not newer than "
                    f"{latest_month.strftime('%b-%y')}"
                )
                continue

            for item in line_items:
                item["Source PDF"] = source_label

            summary_rows.append(
                {
                    "source_pdf": source_label,
                    "sort_date": summary_row[SUMMARY_COLUMNS.index("From Billing Date")],
                    "summary_row": summary_row,
                    "line_items": line_items,
                }
            )

            all_line_items.extend(line_items)

            raw_line_sum = sum(item["Amount"] for item in line_items)
            current_charges = summary_row[SUMMARY_COLUMNS.index("Current Charges")]
            difference = current_charges - raw_line_sum

            validation_rows.append(
                [
                    source_label,
                    summary_row[0],
                    len(line_items),
                    current_charges,
                    raw_line_sum,
                    difference,
                    "OK" if round(difference, 2) == 0 else "CHECK",
                    summary_row[SUMMARY_COLUMNS.index("Review Status")],
                    summary_row[SUMMARY_COLUMNS.index("Review Notes")],
                ]
            )

            review_status = summary_row[SUMMARY_COLUMNS.index("Review Status")]
            print(
                f"ADD: {source_label} | {summary_row[0]} | "
                f"${current_charges:,.2f} | lines={len(line_items)} | {review_status}"
            )

        except Exception as e:
            failed_files.append([source_label, str(e)])
            print(f"FAILED: {source_label} | {e}")

    summary_rows.sort(key=lambda x: (x["sort_date"], x["source_pdf"].lower()))
    all_line_items.sort(key=line_item_sort_key)
    validation_rows.sort(key=lambda row: month_sort_date(row[1]))

    sorted_summary_rows = [item["summary_row"] for item in summary_rows]

    stats = {
        "pdf_files_seen": len(pdf_files),
        "skipped_existing": skipped_existing,
        "skipped_old_or_same_month": skipped_old_or_same_month,
        "new_invoices": len(sorted_summary_rows),
        "failed_files": len(failed_files),
    }

    return sorted_summary_rows, all_line_items, validation_rows, failed_files, stats

# ============================================================
# PROCESS PDFS
# ============================================================


def get_pdf_files(base_folder, year_folders=None):
    """Return invoice PDFs from year folders 2019 through 2026.

    Expected folder structure:
        Invoices\2019
        Invoices\2020
        ...
        Invoices\2026

    Only files with "invoice" in the filename are processed.
    This prevents GCR rate/filing PDFs from being included.
    """
    base_folder = Path(base_folder)

    if not base_folder.exists():
        raise FileNotFoundError(f"Invoice base folder not found:\n{base_folder}")

    if year_folders is None:
        year_folders = DEFAULT_YEAR_FOLDERS

    pdf_files = []
    missing_folders = []

    for year in year_folders:
        folder = base_folder / str(year)

        if not folder.exists():
            missing_folders.append(str(folder))
            continue

        pdf_files.extend(
            pdf_path
            for pdf_path in folder.rglob(PDF_FILE_PATTERN)
            if "invoice" in pdf_path.name.lower()
        )

    if missing_folders:
        print("WARNING: These year folders were not found and were skipped:")
        for folder in missing_folders:
            print(f"  - {folder}")

    return sorted(pdf_files, key=lambda path: str(path).lower())


def process_single_pdf(pdf_path):
    text = extract_pdf_text(pdf_path)
    line_items = parse_invoice_lines(pdf_path, text)
    summary_row = build_summary_row(text, line_items)

    return summary_row, line_items


def process_all_invoices(base_folder, year_folders=None):
    pdf_files = get_pdf_files(base_folder, year_folders)
    base_folder_path = Path(base_folder)

    if not pdf_files:
        years_text = ", ".join(year_folders or DEFAULT_YEAR_FOLDERS)
        raise FileNotFoundError(
            f"No invoice PDF files found in year folders {years_text} under:\n{base_folder}"
        )

    summary_rows = []
    all_line_items = []
    validation_rows = []
    failed_files = []

    for pdf_path in pdf_files:
        try:
            summary_row, line_items = process_single_pdf(pdf_path)

            try:
                source_label = str(pdf_path.relative_to(base_folder_path))
            except ValueError:
                source_label = pdf_path.name

            for item in line_items:
                item["Source PDF"] = source_label

            summary_rows.append(
                {
                    "source_pdf": source_label,
                    "sort_date": summary_row[SUMMARY_COLUMNS.index("From Billing Date")],
                    "summary_row": summary_row,
                    "line_items": line_items,
                }
            )

            all_line_items.extend(line_items)

            raw_line_sum = sum(item["Amount"] for item in line_items)
            current_charges = summary_row[SUMMARY_COLUMNS.index("Current Charges")]
            difference = current_charges - raw_line_sum

            validation_rows.append(
                [
                    source_label,
                    summary_row[0],
                    len(line_items),
                    current_charges,
                    raw_line_sum,
                    difference,
                    "OK" if round(difference, 2) == 0 else "CHECK",
                    summary_row[SUMMARY_COLUMNS.index("Review Status")],
                    summary_row[SUMMARY_COLUMNS.index("Review Notes")],
                ]
            )

            review_status = summary_row[SUMMARY_COLUMNS.index("Review Status")]
            print(
                f"OK: {source_label} | {summary_row[0]} | "
                f"${current_charges:,.2f} | lines={len(line_items)} | {review_status}"
            )

        except Exception as e:
            try:
                failed_label = str(pdf_path.relative_to(base_folder_path))
            except Exception:
                failed_label = pdf_path.name

            failed_files.append([failed_label, str(e)])
            print(f"FAILED: {failed_label} | {e}")

    # Order every worksheet by actual billing month, not alphabetically by filename.
    summary_rows.sort(key=lambda x: (x["sort_date"], x["source_pdf"].lower()))
    all_line_items.sort(key=line_item_sort_key)
    validation_rows.sort(key=lambda row: month_sort_date(row[1]))

    sorted_summary_rows = [item["summary_row"] for item in summary_rows]

    return sorted_summary_rows, all_line_items, validation_rows, failed_files


# ============================================================
# EXCEL WRITER
# ============================================================


def write_excel(output_path, summary_rows, all_line_items, validation_rows, failed_files):
    wb = Workbook()

    ws = wb.active
    ws.title = "Gas Invoice Summary"

    raw_ws = wb.create_sheet("Raw Line Items")
    validation_ws = wb.create_sheet("Validation")

    failed_ws = wb.create_sheet("Failed Files") if failed_files else None

    write_summary_sheet(ws, summary_rows)
    write_raw_line_items(raw_ws, all_line_items)
    write_validation_sheet(validation_ws, validation_rows)

    if failed_ws:
        write_failed_files_sheet(failed_ws, failed_files)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_path)


def write_summary_sheet(ws, summary_rows):
    for col_idx, header in enumerate(SUMMARY_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor="000000")
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, summary_row in enumerate(summary_rows, start=2):
        for col_idx, value in enumerate(summary_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    currency_headers = {
        "Current Charges",
        "1st 0 to 1,500 MMBtu Amount",
        "Next 3,500 MMBtu Amount",
        "All MMBtu over 5,000 MMBtu Amount",
        "Total Cost Based on MMBtu",
        "Customer Charge",
        "Plant Protection Fee",
        "Reimbursement of MGRT 1",
        "Reimbursement of MGRT 2",
        "Reimbursement of MGRT 3",
        "Reimbursement of MGRT 4",
        "Total Reimbursement of MGRT",
        "Street & Alley Fees 1",
        "Street & Alley Fees 2",
        "Street & Alley Fees 3",
        "Street & Alley Fees 4",
        "Total Street & Alley Fees",
        "FIN 48 Tax Refund",
        "Pipeline Safety Fee",
        "Extra Charge Total",
    }

    quantity_headers = {"Billed MCF", "Billed MMBtu", "Billed CCF"}

    rate_headers = {
        "1st 0 to 1,500 MMBtu Rate",
        "Next 3,500 MMBtu Rate",
        "All MMBtu over 5,000 MMBtu Rate",
        "GCR - Industrial Sales ($/MMBtu)",
        "GCR - Transportation ($/MMBtu)",
        "Calculated Rate ($/MMBTU)",
    }

    date_headers = {"From Billing Date", "To Billing Date"}

    for col_idx, header in enumerate(SUMMARY_COLUMNS, start=1):
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if header in currency_headers:
                cell.number_format = "$#,##0.00;[Red]($#,##0.00)"
            elif header in quantity_headers:
                cell.number_format = "#,##0"
            elif header in rate_headers:
                cell.number_format = "0.00000"
            elif header in date_headers:
                cell.number_format = "m/d/yyyy"

    apply_borders(ws)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 55

    for col_idx in range(1, len(SUMMARY_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["E"].width = 30

    total_headers = {
        "Total Cost Based on MMBtu",
        "Total Reimbursement of MGRT",
        "Total Street & Alley Fees",
        "Extra Charge Total",
    }

    for col_idx, header in enumerate(SUMMARY_COLUMNS, start=1):
        if header in total_headers:
            ws.cell(row=1, column=col_idx).fill = PatternFill("solid", fgColor="305496")
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)

    review_col_idx = SUMMARY_COLUMNS.index("Review Status") + 1
    for row_idx in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row_idx, column=review_col_idx)
        status_cell.font = Font(bold=True)

        if status_cell.value == "NEEDS REVIEW":
            status_cell.fill = PatternFill("solid", fgColor="FFC7CE")
        else:
            status_cell.fill = PatternFill("solid", fgColor="C6EFCE")

    ws.column_dimensions[get_column_letter(SUMMARY_COLUMNS.index("Extra Charge Details") + 1)].width = 44
    ws.column_dimensions[get_column_letter(SUMMARY_COLUMNS.index("Review Notes") + 1)].width = 42
    ws.column_dimensions[get_column_letter(SUMMARY_COLUMNS.index("Review Status") + 1)].width = 16


def write_raw_line_items(ws, all_line_items):
    headers = [
        "Source PDF",
        "Line #",
        "Prod Date",
        "Service Account",
        "Customer Name",
        "Description",
        "MCF",
        "MMBtu",
        "Rate",
        "Amount",
        "Review Flag",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, item in enumerate(all_line_items, start=2):
        ws.cell(row=row_idx, column=1, value=item.get("Source PDF", ""))
        ws.cell(row=row_idx, column=2, value=item["Line #"])
        ws.cell(row=row_idx, column=3, value=item["Prod Date"])
        ws.cell(row=row_idx, column=4, value=item["Service Account"])
        ws.cell(row=row_idx, column=5, value=item["Customer Name"])
        ws.cell(row=row_idx, column=6, value=item["Description"])
        ws.cell(row=row_idx, column=7, value=item["MCF"])
        ws.cell(row=row_idx, column=8, value=item["MMBtu"])
        ws.cell(row=row_idx, column=9, value=item["Rate"])
        ws.cell(row=row_idx, column=10, value=item["Amount"])

        description = (item.get("Description") or "").strip()
        if not description or description not in STANDARD_SUMMARY_DESCRIPTIONS:
            ws.cell(row=row_idx, column=11, value="NEEDS REVIEW")
        else:
            ws.cell(row=row_idx, column=11, value="OK")

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=7).number_format = "#,##0"
        ws.cell(row=row_idx, column=8).number_format = "#,##0"
        ws.cell(row=row_idx, column=9).number_format = "0.00000"
        ws.cell(row=row_idx, column=10).number_format = "$#,##0.00;[Red]($#,##0.00)"

        flag_cell = ws.cell(row=row_idx, column=11)
        flag_cell.font = Font(bold=True)
        if flag_cell.value == "NEEDS REVIEW":
            flag_cell.fill = PatternFill("solid", fgColor="FFC7CE")
        else:
            flag_cell.fill = PatternFill("solid", fgColor="C6EFCE")

    apply_borders(ws)
    ws.freeze_panes = "A2"

    widths = {
        "A": 42,
        "B": 10,
        "C": 12,
        "D": 16,
        "E": 30,
        "F": 28,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 16,
        "K": 16,
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def write_validation_sheet(ws, validation_rows):
    headers = [
        "Source PDF",
        "Month",
        "Line Item Count",
        "Summary Current Charges",
        "Raw Line Item Sum",
        "Difference",
        "Math Status",
        "Review Status",
        "Review Notes",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(validation_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=4).number_format = "$#,##0.00;[Red]($#,##0.00)"
        ws.cell(row=row_idx, column=5).number_format = "$#,##0.00;[Red]($#,##0.00)"
        ws.cell(row=row_idx, column=6).number_format = "$#,##0.00;[Red]($#,##0.00)"

        math_status_cell = ws.cell(row=row_idx, column=7)
        if math_status_cell.value == "OK":
            math_status_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        else:
            math_status_cell.fill = PatternFill("solid", fgColor="FFC7CE")

        review_status_cell = ws.cell(row=row_idx, column=8)
        review_status_cell.font = Font(bold=True)
        if review_status_cell.value == "NEEDS REVIEW":
            review_status_cell.fill = PatternFill("solid", fgColor="FFC7CE")
        else:
            review_status_cell.fill = PatternFill("solid", fgColor="C6EFCE")

    apply_borders(ws)
    ws.freeze_panes = "A2"

    widths = {
        "A": 42,
        "B": 12,
        "C": 16,
        "D": 22,
        "E": 22,
        "F": 18,
        "G": 14,
        "H": 16,
        "I": 48,
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def write_failed_files_sheet(ws, failed_files):
    headers = ["Source PDF", "Error"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(failed_files, start=2):
        ws.cell(row=row_idx, column=1, value=row_data[0])
        ws.cell(row=row_idx, column=2, value=row_data[1])

    apply_borders(ws)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 100


def apply_borders(ws):
    thin = Side(border_style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


# ============================================================
# COMMAND-LINE ARGUMENTS
# ============================================================


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Extract Atmos natural-gas invoice PDFs into a multi-sheet Excel "
            "workbook with validation and incremental-update support."
        )
    )
    parser.add_argument(
        "input_folder",
        nargs="?",
        default=str(DEFAULT_INPUT_FOLDER),
        help="Parent folder containing year subfolders such as 2019, 2020, ...",
    )
    parser.add_argument(
        "output_xlsx",
        nargs="?",
        default=str(DEFAULT_OUTPUT_XLSX),
        help="Output Excel workbook path.",
    )
    parser.add_argument(
        "--years",
        nargs="+",
        default=DEFAULT_YEAR_FOLDERS,
        help="Year folders to process, for example: --years 2024 2025 2026",
    )
    parser.add_argument(
        "--full-rebuild",
        action="store_true",
        help="Ignore any existing workbook and rebuild it from the selected invoice years.",
    )
    return parser.parse_args()


# ============================================================
# MAIN
# ============================================================


def main():
    args = parse_args()

    invoice_base_folder = Path(args.input_folder)
    year_folders = [str(year) for year in args.years]
    output_xlsx = Path(args.output_xlsx)

    print("=" * 70)
    print("Atmos Energy Invoice Extractor / Incremental Update")
    print("=" * 70)
    print(f"Invoice base folder: {invoice_base_folder}")
    print(f"Year folders: {', '.join(year_folders)}")
    print(f"Output Excel: {output_xlsx}")

    output_path = output_xlsx

    # Incremental path: if an existing workbook is present, append only newer
    # invoices unless --full-rebuild was requested.
    if output_path.exists() and not args.full_rebuild:
        latest_month, existing_sources, legacy_existing_source_names = read_existing_excel_state(output_path)

        print("-" * 70)
        print("Existing Excel file found.")

        if latest_month:
            print(f"Latest billing month already in Excel: {latest_month.strftime('%b-%y')}")
            print(f"Existing source PDFs already recorded: {len(existing_sources)}")
        else:
            print("Could not find a latest billing month in the existing Excel file.")
            print("Running a full rebuild instead.")
            latest_month = None

        if latest_month:
            summary_rows, all_line_items, validation_rows, failed_files, stats = process_new_invoices(
                invoice_base_folder,
                year_folders,
                latest_month,
                existing_sources,
                legacy_existing_source_names,
            )

            if summary_rows or all_line_items or validation_rows or failed_files:
                append_to_existing_excel(
                    output_path=output_path,
                    summary_rows=summary_rows,
                    all_line_items=all_line_items,
                    validation_rows=validation_rows,
                    failed_files=failed_files,
                )

                print("-" * 70)
                print("Excel updated successfully.")
                print(f"New invoices added: {stats['new_invoices']}")
                print(f"New raw line items added: {len(all_line_items)}")
                print(f"Failed new files: {stats['failed_files']}")
                print(f"Skipped already recorded PDFs: {stats['skipped_existing']}")
                print(f"Skipped old/same-month PDFs: {stats['skipped_old_or_same_month']}")
                print(f"Updated Excel: {output_path}")
                print("=" * 70)
                return

            print("-" * 70)
            print("No new invoice PDFs found after the latest month already in Excel.")
            print(f"Skipped already recorded PDFs: {stats['skipped_existing']}")
            print(f"Skipped old/same-month PDFs: {stats['skipped_old_or_same_month']}")
            print("Nothing was changed.")
            print("=" * 70)
            return

    # Full rebuild path.
    print("-" * 70)
    print("Running full rebuild.")

    pdf_files = get_pdf_files(invoice_base_folder, year_folders)
    base_folder_path = Path(invoice_base_folder)

    print("-" * 70)
    print(f"Invoice PDF files found: {len(pdf_files)}")
    for pdf_path in pdf_files:
        try:
            print(f"  - {pdf_path.relative_to(base_folder_path)}")
        except ValueError:
            print(f"  - {pdf_path.name}")
    print("-" * 70)

    summary_rows, all_line_items, validation_rows, failed_files = process_all_invoices(
        invoice_base_folder,
        year_folders,
    )

    write_excel(
        output_path=output_path,
        summary_rows=summary_rows,
        all_line_items=all_line_items,
        validation_rows=validation_rows,
        failed_files=failed_files,
    )

    print("-" * 70)
    print(f"Successful invoices: {len(summary_rows)}")
    print(f"Failed invoices: {len(failed_files)}")
    print(f"Raw line items: {len(all_line_items)}")
    print("-" * 70)
    print(f"Excel created successfully: {output_path}")
    print("=" * 70)

if __name__ == "__main__":
    main()
