from __future__ import annotations

import re
import argparse
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple
from datetime import datetime, timedelta, date

import pandas as pd
import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# =========================================================
# CONFIG (DEFAULTS)
# =========================================================
PROJECT_ROOT = Path(__file__).resolve().parent
DEFAULT_INPUT = PROJECT_ROOT / "data" / "input"
DEFAULT_OUTPUT = PROJECT_ROOT / "output"

DEBUG = False
AMOUNT_PATTERN = r"(?:-?[0-9,]+\.\d{2}|\([0-9,]+\.\d{2}\))"

# Review tolerance in dollars. Small penny/rounding differences are accepted.
RECONCILIATION_TOLERANCE = 0.05


# =========================================================
# EXACT OUTPUT COLUMNS - YOUR REQUESTED ORDER
# =========================================================
OUTPUT_COLUMNS = [
    "Production Month",
    "From",
    "To",
    "Invoice Date",
    "Power Factor",
    "Load Factor",
    "Actual Demand (KW)",
    "Billing Demand (KW)",
    "4CP Charges Qty (KW)",
    "4CP Charges Rate ($/KW)",
    "4CP Charges ($)",
    "Usage - Actual KWH",
    "UOM",
    "Energy Charge",
    "Nodal Congestion Charge",
    "Market Securitization (Debt) Financing - Default Charge",
    "Prior Period Pass Through Charge",
    "ERCOT Cont Reserve Serv (ECRS)",
    "Firm Fuel Supply Service",
    "Firm Fuel Supply Service - Backbill",
    "Market Securitization - Uplift Charge",
    "TX-ERCOT Admin Fees - CIL",
    "Transmission Charges",
    "Taxes & PUC Assessment Charge",
    "Ancilliary Service Obligation Adjustment",
    "Other Taxes",
    "Bill Total",
]

# Review fields are appended after the requested invoice columns.
REVIEW_COLUMNS = [
    "Review Status",
    "Review Notes",
]

FINAL_OUTPUT_COLUMNS = OUTPUT_COLUMNS + REVIEW_COLUMNS

# Internal-only fields are useful for per-account sheets and troubleshooting.
INTERNAL_COLUMNS = [
    "Account",
    "Provider",
    *FINAL_OUTPUT_COLUMNS,
    "Source File",
]

NUMERIC_COLUMNS = [
    "Actual Demand (KW)",
    "Billing Demand (KW)",
    "4CP Charges Qty (KW)",
    "4CP Charges Rate ($/KW)",
    "4CP Charges ($)",
    "Usage - Actual KWH",
    "Energy Charge",
    "Nodal Congestion Charge",
    "Market Securitization (Debt) Financing - Default Charge",
    "Prior Period Pass Through Charge",
    "ERCOT Cont Reserve Serv (ECRS)",
    "Firm Fuel Supply Service",
    "Firm Fuel Supply Service - Backbill",
    "Market Securitization - Uplift Charge",
    "TX-ERCOT Admin Fees - CIL",
    "Transmission Charges",
    "Taxes & PUC Assessment Charge",
    "Ancilliary Service Obligation Adjustment",
    "Other Taxes",
    "Bill Total",
]

PERCENT_COLUMNS = [
    "Power Factor",
    "Load Factor",
]

DATE_COLUMNS = [
    "Production Month",
    "From",
    "To",
    "Invoice Date",
]


# =========================================================
# ARGUMENTS
# =========================================================
def get_paths() -> Tuple[Path, Path, Path]:
    parser = argparse.ArgumentParser(
        description="Extract energy invoice fields from PDF invoices into Excel/CSV."
    )

    parser.add_argument(
        "input_path",
        nargs="?",
        default=DEFAULT_INPUT,
        help="Input PDF file or folder containing PDF files.",
    )

    parser.add_argument(
        "output_folder",
        nargs="?",
        default=DEFAULT_OUTPUT,
        help="Output folder for Excel and CSV files.",
    )

    args = parser.parse_args()

    input_path = Path(args.input_path)
    output_folder = Path(args.output_folder)

    output_excel = output_folder / "invoice_extracted_table.xlsx"
    output_csv = output_folder / "invoice_extracted_table.csv"

    return input_path, output_excel, output_csv


# =========================================================
# HELPERS
# =========================================================
def normalize_text(text: str) -> str:
    text = text.replace("\u00a0", " ")
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def flatten_text(text: str) -> str:
    t = text.replace("\u00a0", " ")
    t = t.replace("–", "-").replace("—", "-")
    t = re.sub(r"\s+", " ", t)
    return t.strip()


def search_group(pattern: str, text: str, flags: int = 0) -> Optional[str]:
    m = re.search(pattern, text, flags)
    return m.group(1).strip() if m else None


def search_groups(pattern: str, text: str, flags: int = 0) -> Optional[Tuple[str, ...]]:
    m = re.search(pattern, text, flags)
    return m.groups() if m else None


def to_float(value: Optional[str]) -> Optional[float]:
    if value is None:
        return None

    s = str(value).strip().replace(",", "").replace("$", "")
    s = s.replace("(", "-").replace(")", "")

    if s == "":
        return None

    try:
        return float(s)
    except ValueError:
        return None


def nearly_equal(a: Optional[float], b: Optional[float], tol: float = 0.05) -> bool:
    if a is None or b is None:
        return False
    return abs(a - b) <= tol


def convert_invoice_date(date_text: Optional[str]) -> Optional[date]:
    if not date_text:
        return None

    try:
        return datetime.strptime(date_text, "%b %d, %Y").date()
    except ValueError:
        return None


def to_date(date_str: Optional[str]) -> Optional[date]:
    if not date_str:
        return None

    try:
        return datetime.strptime(date_str, "%m/%d/%Y").date()
    except Exception:
        return None


def add_one_day(date_str: Optional[str]) -> Optional[date]:
    d = to_date(date_str)
    if d is None:
        return None
    return d + timedelta(days=1)


def production_month_from_range(
    start_date: Optional[date], end_date: Optional[date]
) -> Optional[date]:
    """
    Return first day of dominant month in the billing period.
    Stored as real date, formatted later in Excel as mmmm-yy.
    """
    if not start_date or not end_date:
        return None

    if end_date < start_date:
        return None

    day_counts: Dict[Tuple[int, int], int] = {}
    current = start_date

    while current <= end_date:
        key = (current.year, current.month)
        day_counts[key] = day_counts.get(key, 0) + 1
        current += timedelta(days=1)

    year, month = max(day_counts, key=day_counts.get)
    return date(year, month, 1)


def extract_label_amount(flat_text: str, label_patterns: List[str]) -> Optional[float]:
    for pat in label_patterns:
        m = re.search(pat, flat_text, flags=re.I)
        if m:
            return to_float(m.group(1))
    return None


def percent_text_to_decimal(value: Optional[str]) -> Optional[float]:
    """
    '95.2%' -> 0.952
    '95.2'  -> 0.952
    """
    if value is None:
        return None

    s = str(value).strip().replace("%", "")
    num = to_float(s)

    if num is None:
        return None

    return num / 100.0


def extract_account_name(text: str, pdf_path: Path) -> str:
    patterns = [
        r"ESI\s*ID[:\s]+([0-9]{10,})",
        r"Esi\s*Id[:\s]+([0-9]{10,})",
        r"Account\s*(?:Number|No\.?)[:\s]+([A-Za-z0-9\-]+)",
        r"Account[:\s]+([A-Za-z0-9\-]+)",
        r"Meter[:\s]+([A-Za-z0-9\-/ ]{3,})",
        r"Service\s*Address[:\s]+(.+?)(?:\n|$)",
    ]

    for pat in patterns:
        m = re.search(pat, text, flags=re.I)
        if m:
            value = m.group(1).strip()
            value = re.sub(r"\s+", " ", value)
            if value:
                return value

    return pdf_path.stem.strip()


def sanitize_sheet_name(name: str) -> str:
    name = str(name).strip()
    if not name:
        name = "Unknown"

    name = re.sub(r"[:\\/*?\[\]]", "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:31] if len(name) > 31 else name


def make_unique_sheet_name(base_name: str, used_names: set[str]) -> str:
    base_name = sanitize_sheet_name(base_name)

    if base_name not in used_names:
        used_names.add(base_name)
        return base_name

    counter = 2

    while True:
        suffix = f"_{counter}"
        trimmed = base_name[: 31 - len(suffix)]
        candidate = f"{trimmed}{suffix}"

        if candidate not in used_names:
            used_names.add(candidate)
            return candidate

        counter += 1


def fmt_money(value: Optional[float]) -> str:
    if value is None:
        return "N/A"
    return f"${value:,.2f}"


def value_or_zero(row: Dict[str, Any], field: str) -> float:
    value = row.get(field)
    return 0.0 if value is None else float(value)


def find_matching_charge_lines(text: str, target_amount: float) -> List[str]:
    """Find source-PDF lines whose trailing amount approximately matches target_amount."""
    if abs(target_amount) <= RECONCILIATION_TOLERANCE:
        return []

    target = abs(target_amount)
    matches: List[str] = []
    skip_labels = (
        "bill total",
        "current charges",
        "total current charges",
        "amount due",
        "previous balance",
        "total amount due",
    )

    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()
        if not line or any(label in line.lower() for label in skip_labels):
            continue

        amount_tokens = re.findall(AMOUNT_PATTERN, line)
        for token in amount_tokens:
            amount = to_float(token)
            if amount is None:
                continue
            if abs(abs(amount) - target) <= RECONCILIATION_TOLERANCE:
                # Keep a compact, human-readable source line.
                matches.append(line[:180])
                break

    # preserve order, remove duplicates
    unique: List[str] = []
    seen = set()
    for item in matches:
        key = item.lower()
        if key not in seen:
            seen.add(key)
            unique.append(item)
    return unique[:3]


def fields_present_but_not_parsed(text: str, row: Dict[str, Any]) -> List[str]:
    """Report specific expected labels that appear in the PDF but failed extraction."""
    checks = {
        "Energy Charge": r"Actual Consumption\s*\*\s*Price",
        "Nodal Congestion Charge": r"Nodal Congestion Charge",
        "Market Securitization (Debt) Financing - Default Charge": r"Market Securitization.*?Default",
        "Prior Period Pass Through Charge": r"Total Prior Period",
        "ERCOT Cont Reserve Serv (ECRS)": r"ERCOT Cont Reserve Serv",
        "Firm Fuel Supply Service": r"Firm Fuel Supply Service(?!\s*-?\s*Backbill)",
        "Firm Fuel Supply Service - Backbill": r"Firm Fuel Supply Service\s*-?\s*Backbill",
        "Market Securitization - Uplift Charge": r"Market Securitization\s*-?\s*Uplift",
        "TX-ERCOT Admin Fees - CIL": r"TX-ERCOT Admin Fees\s*-?\s*CIL",
        "Transmission Charges": r"Total TDSP Pass-Through Charges",
        "Taxes & PUC Assessment Charge": r"Total Taxes and Assessments",
        "Ancilliary Service Obligation Adjustment": r"Ancillary Service(?:s)? Obligation",
        "Bill Total": r"(?:Bill Total|Total Current Charges|Current Charges)",
    }

    missing: List[str] = []
    for field, pattern in checks.items():
        if row.get(field) is None and re.search(pattern, text, flags=re.I | re.S):
            missing.append(field)
    return missing


def build_review(row: Dict[str, Any], text: str, base_notes: List[str]) -> Tuple[str, str]:
    """
    Reconcile Bill Total without double-counting nested detail fields.

    Reconciliation uses these non-overlapping summary buckets:
      Energy Charge
      Transmission Charges
      Taxes & PUC Assessment Charge
      Prior Period Pass Through Charge (or its component details when the total is absent)
      Other Taxes

    4CP is NOT added separately because it is part of Transmission Charges.
    Nodal/Default are NOT added separately because they are part of Energy Charge.
    Prior-period detail charges are NOT added when a prior-period total exists.
    """
    issues = list(base_notes)

    bill_total = row.get("Bill Total")
    energy = row.get("Energy Charge")
    transmission = row.get("Transmission Charges")
    taxes = row.get("Taxes & PUC Assessment Charge")
    prior_total = row.get("Prior Period Pass Through Charge")
    other_taxes = row.get("Other Taxes")

    # If the prior-period summary total is absent, use the parsed detail components as a fallback.
    prior_detail_fields = [
        "ERCOT Cont Reserve Serv (ECRS)",
        "Firm Fuel Supply Service",
        "Firm Fuel Supply Service - Backbill",
        "Market Securitization - Uplift Charge",
        "TX-ERCOT Admin Fees - CIL",
        "Ancilliary Service Obligation Adjustment",
    ]
    prior_detail_values = [row.get(field) for field in prior_detail_fields]
    prior_detail_sum = round(sum(float(v) for v in prior_detail_values if v is not None), 2)
    use_prior_detail = prior_total is None and any(v not in (None, 0, 0.0) for v in prior_detail_values)
    prior_for_recon = prior_detail_sum if use_prior_detail else value_or_zero(row, "Prior Period Pass Through Charge")

    # Critical summary fields that should normally be available for reconciliation.
    critical_missing = [
        field
        for field in ["Energy Charge", "Transmission Charges", "Bill Total"]
        if row.get(field) is None
    ]
    if critical_missing:
        issues.append("Missing required parsed field(s): " + ", ".join(critical_missing))

    source_parse_misses = fields_present_but_not_parsed(text, row)
    if source_parse_misses:
        issues.append("Present in PDF but not parsed: " + ", ".join(source_parse_misses))

    if bill_total is not None:
        parsed_summary = round(
            value_or_zero(row, "Energy Charge")
            + value_or_zero(row, "Transmission Charges")
            + value_or_zero(row, "Taxes & PUC Assessment Charge")
            + prior_for_recon
            + value_or_zero(row, "Other Taxes"),
            2,
        )
        discrepancy = round(float(bill_total) - parsed_summary, 2)

        if abs(discrepancy) > RECONCILIATION_TOLERANCE:
            direction = (
                "Bill Total is higher than the parsed summary"
                if discrepancy > 0
                else "Parsed summary is higher than Bill Total"
            )
            prior_label = "Prior-period detail fallback" if use_prior_detail else "Prior Period"
            issues.append(
                f"Bill total reconciliation mismatch: {direction} by {fmt_money(abs(discrepancy))}. "
                f"Bill Total={fmt_money(float(bill_total))}; Parsed Summary={fmt_money(parsed_summary)}. "
                f"Summary used: Energy={fmt_money(energy)}, Transmission={fmt_money(transmission)}, "
                f"Taxes/PUC={fmt_money(taxes)}, {prior_label}={fmt_money(prior_for_recon)}, "
                f"Other Taxes={fmt_money(other_taxes)}."
            )

            # First, see if discrepancy equals a known nested/detail field.
            detail_candidates = [
                "4CP Charges ($)",
                "Nodal Congestion Charge",
                "Market Securitization (Debt) Financing - Default Charge",
                *prior_detail_fields,
            ]
            matching_fields = []
            for field in detail_candidates:
                val = row.get(field)
                if val is not None and abs(abs(float(val)) - abs(discrepancy)) <= RECONCILIATION_TOLERANCE:
                    matching_fields.append(f"{field}={fmt_money(float(val))}")

            if matching_fields:
                if discrepancy < 0:
                    issues.append(
                        "Possible double-counting: discrepancy matches " + ", ".join(matching_fields) + "."
                    )
                else:
                    issues.append(
                        "Possible omitted/reconciliation component: discrepancy matches "
                        + ", ".join(matching_fields)
                        + "."
                    )
            else:
                source_matches = find_matching_charge_lines(text, discrepancy)
                if source_matches:
                    issues.append(
                        "Possible source line explaining discrepancy: " + " || ".join(source_matches)
                    )
                else:
                    issues.append(
                        f"Unreconciled amount: {fmt_money(abs(discrepancy))}. "
                        "No single parsed/source line matched this amount; review the PDF for an omitted charge, credit, or subtotal."
                    )
        elif not issues:
            return "OK", "Bill Total reconciles to parsed summary within $0.05."

    if issues:
        return "NEEDS REVIEW", " | ".join(issues)
    return "OK", "No validation discrepancies found."


# =========================================================
# TEXT EXTRACTION
# =========================================================
def extract_text_pymupdf(pdf_path: Path) -> str:
    doc = fitz.open(str(pdf_path))
    parts: List[str] = []

    for i, page in enumerate(doc, start=1):
        page_text = page.get_text("text") or ""

        if DEBUG and i == 1:
            print(f"\n--- FIRST PAGE TEXT PREVIEW: {pdf_path.name} ---\n")
            print(page_text[:4500])
            print("\n" + "=" * 100 + "\n")

        parts.append(page_text)

    doc.close()
    return normalize_text("\n".join(parts))


# =========================================================
# PARSER
# =========================================================
def parse_invoice(pdf_path: Path) -> Dict[str, Any]:
    text = extract_text_pymupdf(pdf_path)
    flat = flatten_text(text)

    if not text.strip():
        raise ValueError("No readable text found in PDF")

    row: Dict[str, Any] = {c: None for c in INTERNAL_COLUMNS}
    notes: List[str] = []

    row["Account"] = extract_account_name(text, pdf_path)
    row["Provider"] = "Texas GLO/State Power Program"
    row["UOM"] = "KWH"
    row["Other Taxes"] = 0.0
    row["Firm Fuel Supply Service"] = 0.0
    row["Firm Fuel Supply Service - Backbill"] = 0.0
    row["Ancilliary Service Obligation Adjustment"] = 0.0
    row["Source File"] = pdf_path.name

    # -----------------------------------------------------
    # Invoice Date
    # -----------------------------------------------------
    billing_date_raw = search_group(
        r"Billing Date:\s*([A-Za-z]{3}\s+\d{1,2},\s+\d{4})", text
    ) or search_group(
        r"Account Summary\s*Billing Date:\s*([A-Za-z]{3}\s+\d{1,2},\s+\d{4})",
        text,
        flags=re.S,
    )
    row["Invoice Date"] = convert_invoice_date(billing_date_raw)

    # -----------------------------------------------------
    # Power Factor / Load Factor as decimals
    # -----------------------------------------------------
    power_factor_raw = search_group(r"Power Factor\s+([0-9.]+%?)", text, flags=re.I)
    row["Power Factor"] = percent_text_to_decimal(power_factor_raw)

    load_factor_raw = search_group(r"Load Factor\s+([0-9.]+)\s*%", text, flags=re.I)
    row["Load Factor"] = percent_text_to_decimal(load_factor_raw)

    # -----------------------------------------------------
    # Billing Periods
    # -----------------------------------------------------
    current_period = search_groups(
        r"Current Electric Charges Detail\s*\d+\s*Day Billing Period From\s*([0-9/]+)\s*To\s*([0-9/]+)",
        text,
        flags=re.S | re.I,
    )
    tdsp_period = search_groups(
        r"TDSP Pass-Through Charges\s*From\s*([0-9/]+)\s*To\s*([0-9/]+)",
        text,
        flags=re.S | re.I,
    )

    start_date: Optional[date] = None
    end_date: Optional[date] = None

    if current_period:
        raw_start, raw_end = current_period
        start_date = add_one_day(raw_start)
        end_date = to_date(raw_end)
    elif tdsp_period:
        raw_start, raw_end = tdsp_period
        start_date = add_one_day(raw_start)
        end_date = to_date(raw_end)

    row["From"] = start_date
    row["To"] = end_date
    row["Production Month"] = production_month_from_range(start_date, end_date)

    # -----------------------------------------------------
    # Energy Section
    # -----------------------------------------------------
    fixed_price_amount: Optional[float] = None
    usage_kwh: Optional[float] = None

    acp = search_groups(
        rf"Actual Consumption\s*\*\s*Price\s*([0-9,]+)\s*kWh\s*@\s*\$([0-9.]+)/kWh\s*({AMOUNT_PATTERN})",
        text,
        flags=re.S | re.I,
    )

    if acp:
        kwh, rate, fixed_amt = acp
        usage_kwh = to_float(kwh)
        fixed_price_amount = to_float(fixed_amt)
        row["Usage - Actual KWH"] = usage_kwh

    nodal = to_float(
        search_group(
            rf"Nodal Congestion Charge\s*({AMOUNT_PATTERN})", text, flags=re.S | re.I
        )
    )
    row["Nodal Congestion Charge"] = nodal

    default_charge = extract_label_amount(
        flat,
        [
            rf"Market Securitization\s*\(Debt\)\s*Financing\s*-\s*Default\s*Charge\s*({AMOUNT_PATTERN})",
            rf"Market Securitization\s*\(Debt\)\s*Financing\s*Default\s*[A-Za-z]{{3}}\s*({AMOUNT_PATTERN})",
            rf"Market Securitization\s*\(Debt\)\s*Financing\s*-\s*Default\s*-\s*[A-Za-z]{{3}}\s*({AMOUNT_PATTERN})",
            rf"Market Securitization\s*\(Debt\)\s*Financing.*?Default.*?({AMOUNT_PATTERN})",
        ],
    )
    row["Market Securitization (Debt) Financing - Default Charge"] = default_charge

    energy_total = None
    if (
        fixed_price_amount is not None
        or nodal is not None
        or default_charge is not None
    ):
        energy_total = round(
            (fixed_price_amount or 0.0) + (nodal or 0.0) + (default_charge or 0.0),
            2,
        )
    row["Energy Charge"] = energy_total

    # -----------------------------------------------------
    # Actual Demand / Billing Demand
    # -----------------------------------------------------
    duos = search_groups(
        rf"Distribution\s*Charge\s*\(DUOS\)\s*([0-9,]+)\s*kW\s*@\s*\$([0-9.]+)/kW\s*({AMOUNT_PATTERN})",
        text,
        flags=re.S | re.I,
    )
    if duos:
        row["Actual Demand (KW)"] = to_float(duos[0])

    nuclear = search_groups(
        rf"Nuclear\s*Decommissioning\s*Fee\s*([0-9,]+)\s*kW\s*@\s*\$([0-9.]+)/kW\s*({AMOUNT_PATTERN})",
        text,
        flags=re.S | re.I,
    )
    if nuclear:
        row["Billing Demand (KW)"] = to_float(nuclear[0])

    if row["Billing Demand (KW)"] is None and row["Actual Demand (KW)"] is not None:
        row["Billing Demand (KW)"] = row["Actual Demand (KW)"]

    # -----------------------------------------------------
    # 4CP Charges
    # -----------------------------------------------------
    cp4_exact_rate = None
    cp4 = search_groups(
        rf"Transmission\s*Cost\s*Recov\s*Factor\s*([0-9,]+)\s*kW\s*@\s*\$([0-9.]+)/kW\s*({AMOUNT_PATTERN})",
        text,
        flags=re.S | re.I,
    )

    if cp4:
        qty, rate, charge = cp4
        cp4_qty_val = to_float(qty)
        cp4_exact_rate = to_float(rate)
        cp4_charge_val = to_float(charge)

        row["4CP Charges Qty (KW)"] = cp4_qty_val
        row["4CP Charges Rate ($/KW)"] = cp4_exact_rate
        row["4CP Charges ($)"] = cp4_charge_val

    # -----------------------------------------------------
    # Transmission / Taxes
    # -----------------------------------------------------
    transmission_charges = to_float(
        search_group(
            rf"Total TDSP Pass-Through Charges\s*({AMOUNT_PATTERN})",
            text,
            flags=re.S | re.I,
        )
    )
    row["Transmission Charges"] = transmission_charges

    total_taxes_assessments = to_float(
        search_group(
            rf"Total Taxes and Assessments\s*({AMOUNT_PATTERN})",
            text,
            flags=re.S | re.I,
        )
    )
    row["Taxes & PUC Assessment Charge"] = total_taxes_assessments

    # -----------------------------------------------------
    # Prior Period / ERCOT / Firm Fuel / Uplift / Ancillary
    # -----------------------------------------------------
    prior_period_total = extract_label_amount(
        flat,
        [
            rf"Total Prior Period Pass-Thru Charges\s*({AMOUNT_PATTERN})",
            rf"Total Prior Period Charges\s*({AMOUNT_PATTERN})",
        ],
    )
    row["Prior Period Pass Through Charge"] = prior_period_total

    ecrs = extract_label_amount(
        flat,
        [
            rf"ERCOT Cont Reserve Serv\s*\(ECRS\)\s*({AMOUNT_PATTERN})",
            rf"ERCOT Cont Reserve Serv\s*({AMOUNT_PATTERN})",
        ],
    )
    row["ERCOT Cont Reserve Serv (ECRS)"] = ecrs

    firm_fuel_backbill = extract_label_amount(
        flat,
        [
            rf"Firm Fuel Supply Service\s*-\s*Backbill\s*({AMOUNT_PATTERN})",
            rf"Firm Fuel Supply Service\s*Backbill\s*({AMOUNT_PATTERN})",
        ],
    )
    row["Firm Fuel Supply Service - Backbill"] = (
        firm_fuel_backbill if firm_fuel_backbill is not None else 0.0
    )

    firm_fuel = extract_label_amount(
        flat,
        [
            rf"Firm Fuel Supply Service(?!\s*-\s*Backbill)\s*({AMOUNT_PATTERN})",
        ],
    )
    row["Firm Fuel Supply Service"] = firm_fuel if firm_fuel is not None else 0.0

    uplift = extract_label_amount(
        flat,
        [
            rf"Market Securitization\s*-\s*Uplift\s*({AMOUNT_PATTERN})",
            rf"Market Securitization-\s*Uplift\s*({AMOUNT_PATTERN})",
        ],
    )
    row["Market Securitization - Uplift Charge"] = uplift

    tx_ercot = extract_label_amount(
        flat,
        [
            rf"TX-ERCOT Admin Fees\s*-\s*CIL\s*({AMOUNT_PATTERN})",
        ],
    )
    row["TX-ERCOT Admin Fees - CIL"] = tx_ercot

    ancillary = extract_label_amount(
        flat,
        [
            rf"Ancillary Services Obligation Adj\s*({AMOUNT_PATTERN})",
            rf"Ancillary Service Obligation Adjustment\s*({AMOUNT_PATTERN})",
            rf"Ancillary Services Obligation Adjustment\s*({AMOUNT_PATTERN})",
        ],
    )
    row["Ancilliary Service Obligation Adjustment"] = (
        ancillary if ancillary is not None else 0.0
    )

    # -----------------------------------------------------
    # Bill Total
    # -----------------------------------------------------
    bill_total = extract_label_amount(
        flat,
        [
            rf"Bill Total\s*({AMOUNT_PATTERN})",
            rf"Total Current Charges\s*({AMOUNT_PATTERN})",
            rf"Current Charges\s*({AMOUNT_PATTERN})",
        ],
    )
    row["Bill Total"] = bill_total

    # -----------------------------------------------------
    # Validation / Review Status
    # -----------------------------------------------------
    cp4_qty = row["4CP Charges Qty (KW)"]
    cp4_amt = row["4CP Charges ($)"]

    if cp4_qty is not None and cp4_exact_rate is not None and cp4_amt is not None:
        expected = round(cp4_qty * cp4_exact_rate, 2)
        if not nearly_equal(expected, cp4_amt):
            notes.append(
                f"4CP mismatch: Qty x Rate={fmt_money(expected)}, parsed 4CP charge={fmt_money(cp4_amt)}"
            )

    review_status, review_notes = build_review(row, text, notes)
    row["Review Status"] = review_status
    row["Review Notes"] = review_notes
    return row


# =========================================================
# FILES / DATAFRAME
# =========================================================
def get_pdf_files(input_path: str | Path) -> List[Path]:
    p = Path(input_path)

    if p.is_file() and p.suffix.lower() == ".pdf":
        return [p]

    if p.is_dir():
        return sorted(p.rglob("*.pdf"))

    raise FileNotFoundError(f"Input path not found or invalid: {p}")


def extract_all(input_path: str | Path) -> pd.DataFrame:
    pdf_files = get_pdf_files(input_path)
    rows: List[Dict[str, Any]] = []

    for pdf_file in pdf_files:
        try:
            row = parse_invoice(pdf_file)
            rows.append(row)
        except Exception as e:
            error_row = {c: None for c in INTERNAL_COLUMNS}
            error_row["Account"] = pdf_file.stem
            error_row["Provider"] = "Texas GLO/State Power Program"
            error_row["Source File"] = pdf_file.name
            error_row["Review Status"] = "NEEDS REVIEW"
            error_row["Review Notes"] = f"ERROR while parsing invoice: {e}"
            rows.append(error_row)

    df = pd.DataFrame(rows, columns=INTERNAL_COLUMNS)

    # Enforce data types.
    for col in NUMERIC_COLUMNS + PERCENT_COLUMNS:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    for col in DATE_COLUMNS:
        df[col] = pd.to_datetime(df[col], errors="coerce")

    return df[INTERNAL_COLUMNS]


# =========================================================
# EXCEL FORMATTING
# =========================================================
def apply_excel_formats(excel_path: Path) -> None:
    wb = load_workbook(excel_path)

    money_columns = {
        "4CP Charges ($)",
        "Energy Charge",
        "Nodal Congestion Charge",
        "Market Securitization (Debt) Financing - Default Charge",
        "Prior Period Pass Through Charge",
        "ERCOT Cont Reserve Serv (ECRS)",
        "Firm Fuel Supply Service",
        "Firm Fuel Supply Service - Backbill",
        "Market Securitization - Uplift Charge",
        "TX-ERCOT Admin Fees - CIL",
        "Transmission Charges",
        "Taxes & PUC Assessment Charge",
        "Ancilliary Service Obligation Adjustment",
        "Other Taxes",
        "Bill Total",
    }

    rate_columns = {
        "4CP Charges Rate ($/KW)",
    }

    number_columns = {
        "Actual Demand (KW)",
        "Billing Demand (KW)",
        "4CP Charges Qty (KW)",
        "Usage - Actual KWH",
    }

    percent_columns = set(PERCENT_COLUMNS)
    date_columns = {"From", "To", "Invoice Date"}
    production_month_column = "Production Month"

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for header_cell in ws[1]:
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = ws.cell(row=1, column=cell.column).value

                if header is None or cell.value is None:
                    continue

                if header in money_columns:
                    cell.number_format = "$#,##0.00"
                elif header in rate_columns:
                    cell.number_format = "$0.000000"
                elif header in number_columns:
                    cell.number_format = "#,##0.00"
                elif header in percent_columns:
                    cell.number_format = "0.0%"
                elif header in date_columns:
                    cell.number_format = "m/d/yyyy"
                elif header == production_month_column:
                    cell.number_format = "mmmm-yy"

        # Width adjustment
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter

            for c in col_cells:
                val = "" if c.value is None else str(c.value)
                # Handle multi-line headers
                val_len = max(len(part) for part in val.split("\n"))
                max_len = max(max_len, val_len)

            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 35)

        ws.row_dimensions[1].height = 42


        # Review columns: wrap notes and make NEEDS REVIEW visually obvious.
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        if "Review Notes" in headers:
            review_notes_col = headers["Review Notes"]
            ws.column_dimensions[ws.cell(row=1, column=review_notes_col).column_letter].width = 65
            for cell in ws.iter_cols(min_col=review_notes_col, max_col=review_notes_col, min_row=2):
                for c in cell:
                    c.alignment = Alignment(vertical="top", wrap_text=True)

        if "Review Status" in headers:
            status_col = headers["Review Status"]
            for row_idx in range(2, ws.max_row + 1):
                c = ws.cell(row=row_idx, column=status_col)
                if c.value == "NEEDS REVIEW":
                    c.font = Font(bold=True)

    wb.save(excel_path)


# =========================================================
# SAVE OUTPUTS
# =========================================================
def save_outputs(
    df: pd.DataFrame, excel_path: str | Path, csv_path: str | Path
) -> None:
    excel_path = Path(excel_path)
    csv_path = Path(csv_path)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    df_to_write = df.copy().sort_values(
        by=["Account", "Production Month", "From"], na_position="last"
    )

    # Requested invoice columns first, then Review Status and Review Notes.
    final_output = df_to_write[FINAL_OUTPUT_COLUMNS].copy()
    final_output.to_csv(csv_path, index=False, date_format="%Y-%m-%d")

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        datetime_format="m/d/yyyy",
        date_format="m/d/yyyy",
    ) as writer:
        final_output.to_excel(writer, sheet_name="All Invoices", index=False)

        # Separate review log keeps source/account troubleshooting information available.
        review_log = df_to_write[
            ["Account", "Provider", "Production Month", "Source File", "Review Status", "Review Notes"]
        ].copy()
        review_log.to_excel(writer, sheet_name="Review Log", index=False)

        used_sheet_names: set[str] = {"All Invoices", "Review Log"}
        for account_name, group in df_to_write.groupby("Account", dropna=False):
            safe_account = account_name if str(account_name).strip() else "Unknown"
            sheet_name = make_unique_sheet_name(str(safe_account), used_sheet_names)
            group[FINAL_OUTPUT_COLUMNS].to_excel(writer, sheet_name=sheet_name, index=False)

    apply_excel_formats(excel_path)


# =========================================================
# MAIN
# =========================================================
if __name__ == "__main__":
    INPUT_PATH, OUTPUT_EXCEL, OUTPUT_CSV = get_paths()

    print(f"\nInput path   : {INPUT_PATH}")
    print(f"Excel output : {OUTPUT_EXCEL}")
    print(f"CSV output   : {OUTPUT_CSV}")

    df = extract_all(INPUT_PATH)

    print("\nFINAL TABLE - EXACT OUTPUT ORDER:\n")
    print(df[FINAL_OUTPUT_COLUMNS].to_string(index=False))


    review_df = df[["Source File", "Review Status", "Review Notes"]].copy()
    review_df = review_df[review_df["Review Status"] == "NEEDS REVIEW"]
    if not review_df.empty:
        print("\nINVOICES NEEDING REVIEW:\n")
        print(review_df.to_string(index=False))

    save_outputs(df, OUTPUT_EXCEL, OUTPUT_CSV)

    print(f"\nSaved Excel: {OUTPUT_EXCEL}")
    print(f"Saved CSV:   {OUTPUT_CSV}")
